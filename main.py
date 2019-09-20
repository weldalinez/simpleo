import processes as ps
import time
import statistics
from openpyxl import Workbook,load_workbook

startTime = time.time()
n = 2
w = 2
p = 2
student = 43
question = 5
human_rater = [35, 73, 36, 62, 42, 75, 90, 65, 86, 68, 97, 57, 15, 77, 92, 97, 93, 96, 80, 87, 88, 98, 94, 96, 98, 79,
               79, 81, 70, 74, 37, 81, 37, 77, 90, 73, 42, 96, 87, 94, 88, 63, 72]
no1 = []
excel = []
list_score = []
for count in range(1, student + 1):
    print('===========================================================================================================')
    print("MAHASISWA", count, "\n")
    arr_score = []
    current_score = 0
    score = 0
    doc = ps.read_txt("mahasiswa" + str(count) + ".docx")
    for q in range(0, question):
        kj = 1
        scores = []

        # process the student's answer document
        prep = ps.preprocessing(doc[q])                       #delete repetitive from question, convert to romaji(if needed), filter text

        print('\nJAWABAN ', q + 1)

        key = ps.read_txt("jwbDosen" + str(q + 1) + ".docx")  #read answer key documents
        #for each answer keys (from each questions)
        for x in range(0, len(key)):
            #process answer keys
            prep2 = ps.preprocessing(key[x])                  #delete repetitive from question, convert to romaji(if needed), filter text
            tdmkj = ps.TDMRef(prep2)
            tdms = ps.TDMTest(prep2,prep)
            svdkj = ps.SVD(tdmkj)
            svds = ps.SVD(tdms)
            frobnorm = ps.frobeniusNorm(svdkj,svds)
            scores.append(frobnorm)

            # #print all processes' results
            print('----------------')
            print('KUNCI JAWABAN '+str(kj))
            print('---\npreprocessing dosen\n', prep2)
            # print('---\ntdm kj\n', tdmkj)
            print('---\npreprocessing siswa\n', prep)
            # print('---\ntdm student\n', tdms)
            # print('---\nSVD KJ: ', svdkj, "SVD S: ", svds, "Nilai KJ ke-", kj, ": ", frobnorm)
            # print('---\nSVD S: ', svds)
            # print('---\nNilai KJ ke-', kj, '\t: ', frobnorm)

            kj += 1 #loop for each answer keys
        # # print score of each questions (maximum score from list of each key answers' score)
        # print('\n\nNilai Soal Nomor', q + 1,'\t: ', max(scores), '\n--------------------------')
        #add the score of each questions to list of scores of each students
        arr_score.append(max(scores))
        # excel.append(max(scores))
    # # print students' score (sum of scores from each questions)
    # print('\n\nNILAI MAHASISWA ', count, '\t: ', round(sum(arr_score),2))
    list_score.append(round(sum(arr_score),2))
akurasi = []
for n in range(student):
    acc = round((100 - (((abs(list_score[n] - human_rater[n])) / 100) * 100)),2)
    akurasi.append(acc)
print("===========================================================================================================")
print("Human Rater\t\t\t: ", human_rater)
print("Nilai-nilai Siswa\t: ", list_score)
print("Akurasi\t\t\t\t: ", akurasi)
print("Rata-rata Akurasi\t: ", round(statistics.mean(akurasi),2))
print("Standar Deviasi\t\t: ", round(statistics.stdev(akurasi),2))
print("Program Execution Duration: ", (time.time() - startTime), "seconds")

#create excel file of the data
# wb = Workbook()                                             # Workbook is created
# wb = load_workbook('data.xlsx')
# sheet1 = wb.create_sheet("Test")                  # add_sheet is used to create sheet.
# for n in range(1, 43):
#     for o in range(len(excel)):
#         for m in range(2,6):
#             sheet1.cell(row=n, column=1, value=n)
#             sheet1.cell(row=n, column=m, value=excel[o])
# for n in range(len(human_rater)):
#     sheet1.cell(row=1, column=1, value='Siswa')
#     sheet1.cell(row=1, column=2, value='Human Rater Score')
#     sheet1.cell(row=1, column=3, value='Simple-O Score')
#     sheet1.cell(row=1, column=4, value='Accuracy')
#     sheet1.cell(row=n+2, column=1, value=n)
#     sheet1.cell(row=n+2, column=2, value=human_rater[n])
#     sheet1.cell(row=n+2, column=3, value=list_score[n])
#     sheet1.cell(row=n+2, column=4, value=akurasi[n])
# wb.save('data.xlsx')